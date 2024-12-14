import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import logging
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from logging.handlers import RotatingFileHandler
import os


#----------------CONFIGURANDO MEU LOGGING ----------------------------------------

# Criação de um logger
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Caminho para salvar o arquivo de log
log_dir = os.path.join(os.getcwd(), 'Log')
log_file_path = os.path.join(log_dir, 'Formatar.log')

# Verifica se a pasta Log existe, se não, cria
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# Configuração do handler de log com rotação
handler = RotatingFileHandler(
    log_file_path, 
    maxBytes=2 * 1024 * 1024,  # 2 MB
    backupCount=4,  # Mantém até 5 arquivos de log (original + 4 backups)
    encoding='utf-8'
)

# Configuração do arquivo de log
formatter = logging.Formatter('%(asctime)s - %(filename)s - %(levelname)s - %(message)s', '%Y-%m-%d %H:%M:%S')
handler.setFormatter(formatter)

# Adiciona o handler ao logger
logger.addHandler(handler)

# Configuração do handler para o terminal (Console)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)

# Adiciona o handler de console ao logger
logger.addHandler(console_handler)

#---------------------------------------------------------------------------------


#---- bloco para adiconar a data de hoje no arquivo ----

data = datetime.now()
ano_mes = data.strftime('%Y%m')

#-------------------------------------------------------

#defina o caminho para o arquivo que deseja formatar a baixo

relatorio_path = f'caminho para o arquivo.xlsx'


class FormatadorArquivo:
    @staticmethod
    def formatar():
        try:
            logging.info('> Vamos iniciar o processo de formatação do relatório.')

            # Abrir o arquivo Excel
            wb = openpyxl.load_workbook(relatorio_path)
            ws = wb.active  # Usar a primeira planilha ativa

            # Definir estilos para o cabeçalho
            header_fill = PatternFill(start_color="007858", end_color="007858", fill_type="solid")  # Fdefine a cor de fundo do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")  # Fonte branca e em negrito
            #este bloco define a cor e espessura da borda
            medium_border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )

            logging.info('> vamos aplicar a formatação nos cabeçalhos')
            # Aplicar formatação no cabeçalho (primeira linha)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = medium_border
            logging.info('> Cabeçalhos formatados com sucesso!')

            logging.info('> vamos aplicar a formatação nas linhas')
            # Aplicar bordas para todas as células da planilha
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = medium_border
            logging.info('> Formatação das linhas concluída com sucesso!')

            logging.info('> vamos ajustar as larguras das colunas')
           
            # Ajustar automaticamente a largura das colunas, exceto a coluna F
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)  # Obtém a letra da coluna caso precise ignorar alguma coluna especifica coloque um if col = f: pass ex.

                # Ajustar largura das outras colunas
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.05  # Ajustar com base no tamanho do conteúdo
                ws.column_dimensions[col_letter].width = adjusted_width
            logging.info('> Largura das colunas ajustadas com sucesso!')

            logging.info('> vamos ocultar a exibição das linhas de grade')
            # Ocultar as linhas de grade
            ws.sheet_view.showGridLines = False
            logging.info('> Linhas de grade ocultadas com sucesso!')

            # Salvar o arquivo formatado no local desejado neste caso na mesma pasta
            wb.save(relatorio_path)
            logging.info(f'> Formatação concluída com sucesso para o arquivo {relatorio_path}.')

        
        except Exception as e:
            logging.error(f'> Erro ao formatar o arquivo: {e}')         

def main():

    start = FormatadorArquivo()
    start.formatar()

if __name__ == '__main__':
    main()
